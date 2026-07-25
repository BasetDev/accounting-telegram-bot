"""Export service for generating Excel and PDF reports."""

import os
from typing import List, Dict
from io import BytesIO

from app.config import settings
from app.utils.logger import logger
from app.utils.jdatetime_helper import get_jalali_date, format_amount

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def _ensure_export_dir():
    """Ensure export directory exists."""
    if not os.path.exists(settings.EXPORT_DIR):
        os.makedirs(settings.EXPORT_DIR, exist_ok=True)


def _persian_font_path() -> str:
    """Find a Persian-capable font for PDF generation."""
    # Common Linux font paths
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
        "/usr/share/fonts/opentype/noto/NotoSans-Regular.ttf",
        "/usr/share/fonts/TTF/DejaVuSans.ttf",
    ]
    for path in font_paths:
        if os.path.exists(path):
            return path
    return "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"


async def export_transactions_excel(
    transactions: List[Dict],
    filename: str = None
) -> str:
    """Export transactions to Excel file and return file path."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")
    
    _ensure_export_dir()
    
    if not filename:
        filename = f"transactions_{get_jalali_date().replace('/', '-')}.xlsx"
    
    filepath = os.path.join(settings.EXPORT_DIR, filename)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تراکنش‌های مالی"
    ws.sheet_view.rightToLeft = True
    
    # Styles
    header_font = Font(name="B Nazanin", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_font = Font(name="B Nazanin", size=11)
    cell_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Check if transactions have due dates (debt/receivable)
    has_due_dates = any(t.get("due_jalali_date") for t in transactions)

    # Headers
    headers = ["ردیف", "طرف حساب", "نوع", "مبلغ (تومان)", "دسته‌بندی", "توضیحات", "تاریخ", "ساعت", "وضعیت"]
    if has_due_dates:
        headers.append("سررسید")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    type_names = {
        "income": "💰 درآمد",
        "expense": "💸 هزینه",
        "debt": "📋 بدهی",
        "receivable": "📌 طلب"
    }
    
    for row_idx, txn in enumerate(transactions, 2):
        data = [
            row_idx - 1,
            txn.get("party_name") or "-",
            type_names.get(txn["transaction_type"], txn["transaction_type"]),
            f"{txn["amount"]:,.0f}",
            txn.get("category") or "-",
            txn.get("description") or "-",
            txn["jalali_date"],
            txn["jalali_time"],
            "تسویه شده" if txn.get("is_settled", False) else "جاری"
        ]
        if has_due_dates:
            data.append(txn.get("due_jalali_date") or "-")
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    # Adjust column widths
    from openpyxl.utils import get_column_letter
    column_widths = [6, 20, 14, 18, 16, 30, 14, 10, 12]
    if has_due_dates:
        column_widths.append(14)
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    wb.save(filepath)
    logger.info(f"Excel report exported: {filepath}")
    return filepath


async def export_transactions_pdf(
    transactions: List[Dict],
    filename: str = None
) -> str:
    """Export transactions to PDF file and return file path."""
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is not installed. Install with: pip install reportlab")
    
    _ensure_export_dir()
    
    if not filename:
        filename = f"transactions_{get_jalali_date().replace('/', '-')}.pdf"
    
    filepath = os.path.join(settings.EXPORT_DIR, filename)
    
    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    
    # Try to register Persian font
    try:
        font_path = _persian_font_path()
        pdfmetrics.registerFont(TTFont("Persian", font_path))
        font_name = "Persian"
    except Exception:
        font_name = "Helvetica"
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PersianTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=16,
        alignment=1,  # Center
        spaceAfter=20
    )
    
    # Title
    elements.append(Paragraph("گزارش تراکنش‌های مالی", title_style))
    elements.append(Spacer(1, 12))
    
    # Table data
    has_due_dates = any(t.get("due_jalali_date") for t in transactions)

    type_names = {
        "income": "درآمد",
        "expense": "هزینه",
        "debt": "بدهی",
        "receivable": "طلب"
    }

    headers = ["ردیف", "طرف حساب", "نوع", "مبلغ", "دسته‌بندی", "توضیحات", "تاریخ", "ساعت", "وضعیت"]
    if has_due_dates:
        headers.append("سررسید")
    table_data = [headers]
    
    for idx, txn in enumerate(transactions, 1):
        t_type = type_names.get(txn["transaction_type"], txn["transaction_type"])
        amount = f"{txn["amount"]:,.0f}"
        desc = (txn.get("description") or "")[:30]
        row = [
            str(idx),
            txn.get("party_name") or "-",
            t_type,
            amount,
            txn.get("category") or "-",
            desc,
            txn["jalali_date"],
            txn["jalali_time"],
            "تسویه شده" if txn.get("is_settled", False) else "جاری"
        ]
        if has_due_dates:
            row.append(txn.get("due_jalali_date") or "-")
        table_data.append(row)
    
    # Style
    col_widths = [30, 100, 50, 80, 70, 140, 65, 50, 60]
    if has_due_dates:
        col_widths.append(65)
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    
    elements.append(table)
    
    # Summary - adapt based on transaction types present
    summary_style = ParagraphStyle(
        "PersianSummary",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=11,
        spaceBefore=20
    )

    elements.append(Spacer(1, 20))

    txn_types = set(t["transaction_type"] for t in transactions) if transactions else set()

    if txn_types == {"income"} or txn_types == {"expense"} or txn_types == {"income", "expense"}:
        total_income = sum(t["amount"] for t in transactions if t["transaction_type"] == "income")
        total_expense = sum(t["amount"] for t in transactions if t["transaction_type"] == "expense")
        elements.append(Paragraph(f"مجموع درآمد: {total_income:,.0f} تومان", summary_style))
        elements.append(Paragraph(f"مجموع هزینه: {total_expense:,.0f} تومان", summary_style))
    elif "debt" in txn_types:
        total_amount = sum(t["amount"] for t in transactions)
        settled_count = sum(1 for t in transactions if t.get("is_settled"))
        active_count = sum(1 for t in transactions if not t.get("is_settled"))
        elements.append(Paragraph(f"مجموع مبلغ: {total_amount:,.0f} تومان", summary_style))
        elements.append(Paragraph(f"بدهی‌های فعال: {active_count} مورد", summary_style))
        elements.append(Paragraph(f"تسویه شده: {settled_count} مورد", summary_style))
    elif "receivable" in txn_types:
        total_amount = sum(t["amount"] for t in transactions)
        settled_count = sum(1 for t in transactions if t.get("is_settled"))
        active_count = sum(1 for t in transactions if not t.get("is_settled"))
        elements.append(Paragraph(f"مجموع مبلغ: {total_amount:,.0f} تومان", summary_style))
        elements.append(Paragraph(f"طلب‌های فعال: {active_count} مورد", summary_style))
        elements.append(Paragraph(f"وصول شده: {settled_count} مورد", summary_style))
    else:
        total_income = sum(t["amount"] for t in transactions if t["transaction_type"] == "income")
        total_expense = sum(t["amount"] for t in transactions if t["transaction_type"] == "expense")
        if total_income > 0:
            elements.append(Paragraph(f"مجموع درآمد: {total_income:,.0f} تومان", summary_style))
        if total_expense > 0:
            elements.append(Paragraph(f"مجموع هزینه: {total_expense:,.0f} تومان", summary_style))

    elements.append(Paragraph(f"تعداد تراکنش‌ها: {len(transactions)}", summary_style))
    
    doc.build(elements)
    logger.info(f"PDF report exported: {filepath}")
    return filepath